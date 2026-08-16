#!/usr/bin/env python3
"""Import the reviewed WÖk Master Items workbook into public web metadata.

The workbook remains the source of truth. This importer deliberately does not
derive a score from empty cells or publish an automatic decision.
"""
from __future__ import annotations

import hashlib
import html
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
PUBLIC_WORKBOOK = ROOT / "data/master-register/WOeK_Masterregister_v1.4_FINAL_2026-08-16.xlsx"
SOURCE = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else PUBLIC_WORKBOOK
ITEM_SHEET = "01_Item_Register"
HEADER_ROW = 4


def digest(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def records_from_sheet(workbook, sheet_name: str, header_row: int = HEADER_ROW) -> list[dict[str, object]]:
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "").strip() for cell in sheet[header_row]]
    records: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        record = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        if str(record.get("WOK_ID") or record.get("WÖK_ID") or "").strip():
            records.append(record)
    return records


def esc(value: object) -> str:
    return html.escape("" if value is None else str(value), quote=True)


def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"Missing XLSX: {SOURCE}")
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    items = records_from_sheet(workbook, ITEM_SHEET)
    if len(items) != 621:
        raise SystemExit(f"Expected 621 WÖk IDs in {ITEM_SHEET}, found {len(items)}")

    source_hash = digest(SOURCE)
    public_url = "/assets/downloads/woek-register/v1.4/WOeK_Masterregister_v1.4_FINAL_2026-08-16.xlsx"
    payload = {
        "sourceFile": "data/master-register/WOeK_Masterregister_v1.4_FINAL_2026-08-16.xlsx",
        "originalFileUrl": public_url,
        "sourceHash": source_hash,
        "sourceVersion": "v1.4",
        "importVersion": "2026.4-import",
        "liveReferenceVersion": "2026.4-live-reference",
        "reviewStatus": "geprüft",
        "items": items,
    }
    for path in [ROOT / "public/data/woek-ids.json", ROOT / "src/data/woek-ids.json"]:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")

    source_count = len(records_from_sheet(workbook, "07_Quellenkatalog"))
    rules_count = len(records_from_sheet(workbook, "03_Scoring_Rules"))
    document = f"""<!doctype html>
<html lang="de"><head>
  <meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
  <title>WÖk-Masterregister v1.4 | Wirkungsökonomie</title>
  <meta name="description" content="Geprüftes Arbeits- und Governance-Register mit 621 WÖk-IDs, 28 Scoring-Regeln und dokumentiertem Prüfstatus.">
  <link rel="canonical" href="https://wirkungsoekonomie.de/bibliothek/woek-master-items-register/">
  <link rel="stylesheet" href="../../assets/css/style.css">
</head><body><main class="reference-work" data-pagefind-body><article class="article-shell">
  <nav class="breadcrumb"><a href="../../bibliothek/">Bibliothek</a> / WÖk-Masterregister v1.4</nav>
  <h1>WÖk-Masterregister v1.4</h1>
  <p class="lead">Führendes technisches Register für WÖk-IDs, Bewertungsregeln, Quellen und Prüfstatus.</p>
  <p><a class="button" href="{public_url}">XLSX herunterladen</a> <a class="button secondary" href="../../woek-id-register/">WÖk-ID Register durchsuchen</a></p>
  <section class="callout"><h2>Einordnung</h2><p>v1.4 führt Itemdaten, Scoring-Regeln, Schwellen- und Benchmarkstatus, Quellenfunktionen und Prüfprotokolle zusammen. Leere Eingaben bleiben unbewertet. FINAL bezeichnet die kanonische Registerfassung, nicht die abgeschlossene Validierung jeder Schwelle.</p></section>
  <section><h2>Registerumfang</h2><dl><dt>WÖk-IDs</dt><dd>{len(items)}</dd><dt>Scoring-Regeln</dt><dd>{rules_count}</dd><dt>Quellenkatalog</dt><dd>{source_count}</dd><dt>Stand</dt><dd>16. August 2026</dd><dt>Quelldatei</dt><dd>WOeK_Masterregister_v1.4_FINAL_2026-08-16.xlsx</dd><dt>SHA-256</dt><dd>{esc(source_hash)}</dd></dl></section>
  <section><h2>Versionshinweis</h2><p>v1.4 ersetzt v1.3 als führende technische Registerquelle.</p></section>
</article></main></body></html>"""
    output = ROOT / "bibliothek/woek-master-items-register/index.html"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(document, encoding="utf-8")

    imports_path = ROOT / "public/data/workpaper-imports.json"
    if imports_path.exists():
        imports = json.loads(imports_path.read_text(encoding="utf-8"))
        documents = imports.setdefault("documents", [])
        record = next((entry for entry in documents if entry.get("slug") == "woek-master-items-register"), None)
        if record is None:
            record = {"slug": "woek-master-items-register"}
            documents.append(record)
        record.update({
            "source": payload["sourceFile"], "originalName": SOURCE.name, "originalUrl": public_url,
            "status": "führendes-referenzregister", "reviewStatus": "geprüft", "sourceHash": source_hash,
            "blocks": len(items), "issues": [], "updatedAt": datetime.now(timezone.utc).isoformat(),
        })
        imports_path.write_text(json.dumps(imports, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(items)} WÖk-ID rows from leading Master Register v1.4.")


if __name__ == "__main__":
    main()
