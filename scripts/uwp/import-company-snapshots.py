#!/usr/bin/env python3
"""Build versioned company snapshots for the UWP-100 beta.

The public page must not scrape live provider data. This importer fetches the
official SBTi target-dashboard Excel file into a static snapshot and calculates
only source-bound partial profiles. It does not create a full company impact
score and never treats missing SBTi matches as negative climate performance.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys
import tempfile
import unicodedata
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data" / "uwp"
SNAPSHOT_DIR = DATA_DIR / "snapshots"
UNIVERSE_PATH = DATA_DIR / "company-universe.uwp100.json"
SBTI_URL = "https://files.sciencebasedtargets.org/production/files/companies-excel.xlsx"
METHOD_VERSION = "uwp-100-beta-0.2"


COMPANY_NAMES = [
    "Adidas", "Airbus", "Allianz", "BASF", "Bayer", "Beiersdorf", "BMW", "Brenntag", "Commerzbank", "Continental",
    "Covestro", "Daimler Truck", "Deutsche Bank", "Deutsche Börse", "DHL Group", "Deutsche Telekom", "E.ON", "Fresenius", "Hannover Rück", "Heidelberg Materials",
    "Henkel", "Hochtief", "Infineon", "Mercedes-Benz Group", "Merck KGaA", "MTU Aero Engines", "Munich Re", "Porsche AG", "Porsche Automobil Holding", "Qiagen",
    "Rheinmetall", "RWE", "SAP", "Sartorius", "Siemens", "Siemens Energy", "Siemens Healthineers", "Symrise", "Volkswagen", "Vonovia",
    "Zalando", "1&1", "AIXTRON", "Aroundtown", "Aurubis", "Bechtle", "Bilfinger", "Carl Zeiss Meditec", "CTS Eventim", "Delivery Hero",
    "Dermapharm", "Dürr", "Elmos Semiconductor", "Evotec", "Fielmann", "Fraport", "Freenet", "Fuchs", "GEA", "Gerresheimer",
    "HelloFresh", "Hensoldt", "Hugo Boss", "Jungheinrich", "K+S", "KION", "Knorr-Bremse", "Krones", "Lanxess", "LEG Immobilien",
    "Lufthansa", "Nemetschek", "Nordex", "PUMA", "Rational", "Redcare Pharmacy", "RTL Group", "Scout24", "Siltronic", "Ströer",
    "SÜSS MicroTec", "Talanx", "TeamViewer", "Thyssenkrupp", "Traton", "United Internet", "Wacker Chemie", "Wacker Neuson", "Amadeus Fire", "ATOSS Software",
    "Basler", "BayWa", "Befesa", "Borussia Dortmund", "Cancom", "Ceconomy", "Deutsche EuroShop", "Deutsche Pfandbriefbank", "Drägerwerk", "Eckert & Ziegler",
]

ALIASES = {
    "BMW": ["BMW Group"],
    "Infineon": ["Infineon Technologies AG"],
    "Mercedes-Benz Group": ["Mercedes-Benz AG"],
    "Porsche AG": ["Dr. Ing. h.c. F. Porsche AG"],
    "Qiagen": ["QIAGEN N.V."],
    "Thyssenkrupp": ["thyssenkrupp AG"],
}

BLOCKED_PROVIDER_MATCHES = {
    ("Allianz", "Allianz Investment Management SE"),
    ("BayWa", "Baywa Global Produce"),
    ("Fresenius", "Fresenius Medical Care AG"),
}


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def normalize(value: str | None) -> str:
    if not value:
        return ""
    text = unicodedata.normalize("NFD", value.casefold())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(
        r"\b(aktiengesellschaft|ag|se|sa|plc|group|holding|holdings|kgaa|gmbh|co|kg|inc|ltd|limited|corporation|corp|company|the)\b",
        " ",
        text,
    )
    return re.sub(r"\s+", " ", text).strip()


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def build_universe() -> dict[str, Any]:
    companies = []
    for index, name in enumerate(COMPANY_NAMES):
        sector = "Large Cap / DAX-Umfeld" if index < 40 else "Mid Cap / MDAX-TecDAX-Umfeld" if index < 80 else "Small Cap / SDAX-Umfeld"
        companies.append(
            {
                "company_id": f"uwp-{index + 1:03d}",
                "name": name,
                "country": "Deutschland",
                "sector": sector,
                "universe": "UWP-100 Beta",
                "status": "kuratiertes Beta-Universum, nicht offizieller Index",
            }
        )
    return {
        "universe_id": "uwp-100-beta",
        "title": "UWP-100 Beta",
        "description": "100 vorselektierte Unternehmen aus dem deutschen Börsenumfeld.",
        "method": "Kuratiertes Beta-Universum aus DAX-/MDAX-/TecDAX-/SDAX- bzw. HDAX-Umfeld; kein offizieller Index.",
        "created_at": "2026-06-09",
        "updated_at": utc_now(),
        "source_note": "Beta-Auswahlliste der Wirkungsökonomie. Unternehmenskennzahlen werden nur aus öffentlichen, versionierten Snapshots ergänzt.",
        "companies": companies,
    }


def download_sbti(target: Path) -> None:
    request = urllib.request.Request(SBTI_URL, headers={"User-Agent": "wirkungsoekonomie-uwp-snapshot-importer/0.2"})
    with urllib.request.urlopen(request, timeout=90) as response:
        target.write_bytes(response.read())


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_sbti_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Data"]
    headers = [str(cell) for cell in next(worksheet.iter_rows(values_only=True))]
    rows = []
    for row in worksheet.iter_rows(values_only=True):
        item = dict(zip(headers, row))
        item["_normalized_name"] = normalize(str(item.get("company_name") or ""))
        rows.append(item)
    return rows


def find_match(company: dict[str, Any], rows: list[dict[str, Any]]) -> tuple[dict[str, Any] | None, str]:
    names = [company["name"], *ALIASES.get(company["name"], [])]
    normalized_names = {normalize(name) for name in names if normalize(name)}
    candidates = []
    for row in rows:
        row_name = row["_normalized_name"]
        if not row_name:
            continue
        if row_name in normalized_names:
            candidates.append((3, row))
        elif any(row_name.startswith(name + " ") for name in normalized_names if len(name) >= 4):
            candidates.append((2, row))
        elif any(name and name in row_name.split() for name in normalized_names if len(name) >= 6):
            candidates.append((1, row))
    if not candidates:
        return None, "no_safe_match"
    candidates.sort(key=lambda item: (item[0], item[1].get("location") == "Germany"), reverse=True)
    selected = candidates[0][1]
    if (company["name"], str(selected.get("company_name") or "")) in BLOCKED_PROVIDER_MATCHES:
        return None, "subsidiary_match_blocked"
    return selected, "safe_name_match"


def status_score(status: str | None) -> int:
    value = (status or "").casefold()
    if "targets set" in value:
        return 72
    if "committed" in value:
        return 50
    if "removed" in value:
        return 25
    return 0


def classification_bonus(classification: str | None) -> int:
    value = (classification or "").casefold()
    if "1.5" in value:
        return 10
    if "well-below" in value:
        return 5
    return 0


def clamp(value: int) -> int:
    return max(0, min(100, value))


def score_profile(company: dict[str, Any], match: dict[str, Any] | None) -> dict[str, Any]:
    if not match:
        return {
            "company_id": company["company_id"],
            "company_name": company["name"],
            "status": "Quellenprofil berechnet; SBTi-Snapshot ohne sicheren Treffer",
            "overall_score": None,
            "mensch_score": None,
            "planet_score": None,
            "demokratie_score": None,
            "transformation_score": None,
            "data_quality_score": 32,
            "coverage": 0.22,
            "source_count": 1,
            "observation_count": 0,
            "red_lines": [],
            "interpretation": (
                "Im vorhandenen Datenstand ist das Unternehmen im UWP-100-Universum angelegt. "
                "Der SBTi-Snapshot liefert keinen sicheren Unternehmensmatch. Daraus folgt keine negative Klimabewertung, "
                "sondern eine offene Prüffrage: Welche Berichte, CSRD-/ESRS-Daten, Taxonomie-Kennzahlen und Zielpfade sind versioniert verfügbar?"
            ),
            "calculation_note": "Berechnet wurde nur ein Quellen- und Datenreifeprofil; kein Wirkungs- oder Finanzrating.",
        }
    near_score = status_score(str(match.get("near_term_status") or ""))
    net_zero_score = status_score(str(match.get("net_zero_status") or ""))
    bonus = classification_bonus(str(match.get("near_term_target_classification") or ""))
    planet = clamp(near_score + bonus)
    transformation = clamp(round((near_score * 0.55) + (net_zero_score * 0.35) + bonus))
    quality = 58
    if match.get("isin"):
        quality += 8
    if match.get("lei"):
        quality += 8
    if match.get("full_target_language"):
        quality += 12
    if match.get("target_classification_long"):
        quality += 6
    observations = sum(1 for key in ("near_term_status", "near_term_target_classification", "near_term_target_year", "long_term_status", "net_zero_status", "full_target_language") if match.get(key))
    return {
        "company_id": company["company_id"],
        "company_name": company["name"],
        "status": "Teilprofil berechnet: SBTi-Zielstatus importiert",
        "overall_score": None,
        "mensch_score": None,
        "planet_score": planet,
        "demokratie_score": None,
        "transformation_score": transformation,
        "data_quality_score": clamp(quality),
        "coverage": round(min(0.55, 0.28 + observations * 0.045), 2),
        "source_count": 1,
        "observation_count": observations,
        "red_lines": [],
        "interpretation": (
            f"Im vorhandenen Datenstand liegt ein öffentlicher SBTi-Match vor: {match.get('company_name')}. "
            "Rechnerisch fließen nur Zielstatus, Zielklassifikation und vorhandene Identifier in das Planet-/Transformations-Teilprofil ein. "
            "Das ist kein Nachweis tatsächlicher Emissionsreduktion und kein vollständiger UWP-Gesamtwert."
        ),
        "calculation_note": "Teilberechnung aus dem offiziellen SBTi-Target-Dashboard; Mensch und Demokratie benötigen weitere Quellen.",
    }


def build_snapshot(args: argparse.Namespace) -> Path:
    universe = build_universe()
    write_json(UNIVERSE_PATH, universe)
    with tempfile.TemporaryDirectory() as temp_dir:
        xlsx_path = Path(temp_dir) / "sbti-companies.xlsx"
        download_sbti(xlsx_path)
        rows = load_sbti_rows(xlsx_path)
        profiles = []
        observations = []
        for company in universe["companies"]:
            match, match_method = find_match(company, rows)
            profile = score_profile(company, match)
            profile["match_method"] = match_method
            if match:
                profile["matched_provider_name"] = match.get("company_name")
                profile["isin"] = match.get("isin") or None
                profile["lei"] = match.get("lei") or None
                profile["provider_sector"] = match.get("sector") or None
                profile["provider_location"] = match.get("location") or None
                for key in ("near_term_status", "near_term_target_classification", "near_term_target_year", "long_term_status", "long_term_target_classification", "long_term_target_year", "net_zero_status", "net_zero_year", "target_classification_long"):
                    value = match.get(key)
                    if value is not None and value != "":
                        observations.append(
                            {
                                "observation_id": f"obs-{hashlib.sha1(f'{company['company_id']}|{key}|{value}'.encode('utf-8')).hexdigest()[:16]}",
                                "company_id": company["company_id"],
                                "indicator_id": f"sbti_{key}",
                                "dimension": "planet" if "term" in key or "target" in key else "transformation",
                                "raw_value": value,
                                "unit": "qualitative",
                                "source_id": "sbti-target-dashboard",
                                "extraction_method": "snapshot_xlsx_import",
                                "confidence": "medium" if match_method == "safe_name_match" else "low",
                                "data_quality": "B",
                            }
                        )
            profiles.append(profile)
        snapshot = {
            "snapshot_id": "uwp-100.sbti.latest",
            "universe_id": universe["universe_id"],
            "title": "UWP-100 Beta SBTi Snapshot",
            "method_version": METHOD_VERSION,
            "imported_at": utc_now(),
            "provider_id": "sbti",
            "provider_title": "Science Based Targets initiative Target Dashboard",
            "source_url": SBTI_URL,
            "license": "Public dashboard download; reuse requires provider terms review before downstream publication beyond source-bound display.",
            "file_hash_sha256": file_hash(xlsx_path),
            "score_ready": False,
            "score_note": "Teilprofile berechnet. Kein Gesamtwert, kein ESG-Rating, keine Investmentempfehlung und keine Aussage über tatsächliche Emissionsreduktion ohne zusätzliche Quellen.",
            "company_count": len(universe["companies"]),
            "matched_company_count": sum(1 for profile in profiles if profile.get("matched_provider_name")),
            "observation_count": len(observations),
            "companies": universe["companies"],
            "profiles": profiles,
            "observations": observations,
            "sources": [
                {
                    "source_id": "sbti-target-dashboard",
                    "title": "Companies taking action / Target Dashboard",
                    "provider": "Science Based Targets initiative",
                    "url": SBTI_URL,
                    "retrieved_at": utc_now(),
                    "used_for": "Planet-/Transformations-Teilprofil: Zielstatus, Klassifikation, Zieljahr, Net-Zero-Status.",
                    "limits": "SBTi-Ziele sind Ziel- und Validierungsdaten. Sie ersetzen keine Messung tatsächlicher Wirkung, Scope-Emissionen, Lieferkettenwirkung oder Kontroversenprüfung.",
                }
            ],
        }
        path = SNAPSHOT_DIR / "uwp-100.sbti.latest.json"
        write_json(path, snapshot)
        return path


def validate() -> None:
    universe = read_json(UNIVERSE_PATH)
    if len(universe.get("companies", [])) != 100:
        raise SystemExit("UWP universe must contain exactly 100 companies")
    snapshot_path = SNAPSHOT_DIR / "uwp-100.sbti.latest.json"
    snapshot = read_json(snapshot_path)
    if snapshot.get("company_count") != 100:
        raise SystemExit("UWP snapshot must contain exactly 100 companies")
    for profile in snapshot.get("profiles", []):
        if profile.get("overall_score") is not None:
            raise SystemExit("UWP beta snapshot must not publish overall scores")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("command", nargs="?", choices=["sbti", "validate"], default="sbti")
    args = parser.parse_args()
    if args.command == "validate":
        validate()
        print("UWP snapshots valid")
        return
    path = build_snapshot(args)
    print(path.relative_to(ROOT))


if __name__ == "__main__":
    main()
