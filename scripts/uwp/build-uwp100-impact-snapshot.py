#!/usr/bin/env python3
"""Build a source-bound UWP-100 impact-data snapshot.

This importer reads the existing public UWP-100 universe and does not overwrite
company IDs. It imports the official SBTi target dashboard for all 100 companies
where a safe provider match is possible, merges the manually sourced pilot data
from the provided ZIP (SAP, Siemens, BASF), and renders explicit E gaps for
core UWP fields that are not yet machine-sourced.

It does not calculate company scores.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import tempfile
import time
import unicodedata
import urllib.error
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data" / "uwp"
SNAPSHOT_DIR = DATA_DIR / "snapshots"
UNIVERSE_PATH = DATA_DIR / "company-universe.uwp100.json"
SBTI_URL = "https://files.sciencebasedtargets.org/production/files/companies-excel.xlsx"
WIKIRATE_BASE_URL = "https://wikirate.org"
METHOD_VERSION = "uwp-100-impact-data-0.1"
FETCH_CACHE: dict[str, dict[str, Any] | None] = {}
TEXT_FETCH_CACHE: dict[str, str | None] = {}
LAST_FETCH_AT = 0.0

SBTI_FIELDS = [
    ("near_term_status", "Planet", "Klima/Transformationspfad", "SBTi Near-Term Zielstatus", "Status", "A"),
    ("near_term_target_classification", "Planet", "Klima/Transformationspfad", "SBTi Near-Term Zielklassifikation", "Klassifikation", "A"),
    ("near_term_target_year", "Planet", "Klima/Transformationspfad", "SBTi Near-Term Zieljahr", "Jahr", "A"),
    ("long_term_status", "Transformation", "Net-Zero/Transformation", "SBTi Long-Term Zielstatus", "Status", "A"),
    ("long_term_target_classification", "Transformation", "Net-Zero/Transformation", "SBTi Long-Term Zielklassifikation", "Klassifikation", "A"),
    ("long_term_target_year", "Transformation", "Net-Zero/Transformation", "SBTi Long-Term Zieljahr", "Jahr", "A"),
    ("net_zero_status", "Transformation", "Net-Zero/Transformation", "SBTi Net-Zero Status", "Status", "A"),
    ("net_zero_year", "Transformation", "Net-Zero/Transformation", "SBTi Net-Zero Zieljahr", "Jahr", "A"),
    ("ba15_status", "Planet", "Klima/Transformationspfad", "SBTi Business Ambition 1.5C Status", "Status", "B"),
    ("ba15_date", "Planet", "Klima/Transformationspfad", "SBTi Business Ambition 1.5C Datum", "Datum", "B"),
    ("target_classification_long", "Planet", "Klima/Transformationspfad", "SBTi Zielklassifikation Langtext", "Text", "B"),
    ("full_target_language", "Planet", "Klima/Transformationspfad", "SBTi Zieltext", "Text", "B"),
    ("date_updated", "Datenqualität", "Quellenstand", "SBTi Datenstand", "Datum", "A"),
]

CORE_GAP_FIELDS = [
    ("scope_1_2_ghg_absolute", "Planet", "Klima", "Scope 1+2 absolut", "t CO2e"),
    ("scope_3_ghg_absolute", "Planet", "Lieferkette/Klima", "Scope 3 absolut", "t CO2e"),
    ("taxonomy_aligned_revenue_share", "Transformation", "EU-Taxonomie", "Taxonomie-konformer Umsatzanteil", "%"),
    ("renewable_electricity_share", "Planet", "Energie", "Erneuerbarer Strom", "%"),
    ("women_management_share", "Mensch", "Diversität/Teilhabe", "Frauenanteil Management", "%"),
    ("ltifr", "Mensch", "Arbeitssicherheit", "Lost Time Injury Frequency Rate", "je Mio. Arbeitsstunden"),
    ("cbcr_tax_transparency", "Demokratie", "Steuertransparenz", "Country-by-Country Reporting / Steuertransparenz", "Status"),
    ("lobbying_transparency", "Demokratie", "Lobbying/Einflussnahme", "Lobbying-Transparenz", "Status"),
    ("whistleblowing_cases", "Demokratie", "Compliance/Whistleblowing", "Whistleblowing-Fälle und Abhilfe", "Anzahl/Status"),
    ("data_privacy_incidents", "Demokratie", "Datenschutz", "Datenschutzverstöße / Bußgelder", "Anzahl/Status"),
]

WIKIRATE_METRICS = [
    (
        "Global_Reporting_Initiative+Direct_greenhouse_gas_GHG_emissions_Scope_1_GRI_305_1_a_formerly_G4_EN15_a",
        "wikirate_gri_scope1_ghg",
        "Planet",
        "Klima/Emissionen",
        "GRI 305-1 Scope 1 GHG emissions",
        "tonnes CO2 equivalent",
    ),
    (
        "Global_Reporting_Initiative+Indirect_greenhouse_gas_GHG_emissions_Scope_2_GRI_305_2_formerly_G4_EN16_a",
        "wikirate_gri_scope2_ghg",
        "Planet",
        "Klima/Emissionen",
        "GRI 305-2 Scope 2 GHG emissions",
        "tonnes CO2 equivalent",
    ),
    (
        "Global_Reporting_Initiative+Indirect_greenhouse_gas_GHG_emissions_Scope_3_GRI_305_3_formerly_G4_EN17_a",
        "wikirate_gri_scope3_ghg",
        "Planet",
        "Lieferkette/Klima",
        "GRI 305-3 Scope 3 GHG emissions",
        "tonnes CO2 equivalent",
    ),
    (
        "Global_Reporting_Initiative+Total_energy_consumption_GRI_302_1_e_formerly_G4_EN3_e",
        "wikirate_gri_energy_consumption",
        "Planet",
        "Energie",
        "GRI 302-1 Total energy consumption",
        "Gigajoules",
    ),
    (
        "Global_Reporting_Initiative+Total_Waste_Generated_GRI_306_2_formerly_G4_EN23_a",
        "wikirate_gri_total_waste",
        "Planet",
        "Abfall/Ressourcen",
        "GRI 306-2 Total waste generated",
        "tonnes",
    ),
    (
        "Global_Reporting_Initiative+Employees_according_to_gender_G4_LA12",
        "wikirate_gri_female_employees_share",
        "Mensch",
        "Diversität/Belegschaft",
        "GRI G4-LA12 female employees share",
        "%",
    ),
    (
        "Global_Reporting_Initiative+Injury_rate_GRI_403_9_a_formerly_G4_LA6_a",
        "wikirate_gri_injury_rate",
        "Mensch",
        "Arbeitssicherheit",
        "GRI 403-9 injury rate",
        "injuries per million hours worked",
    ),
    (
        "Global_Reporting_Initiative+Worker_fatalities_GRI_403_9_formerly_G4_LA6_a",
        "wikirate_gri_worker_fatalities",
        "Mensch",
        "Arbeitssicherheit",
        "GRI 403-9 worker fatalities",
        "count",
    ),
    (
        "Global_Reporting_Initiative+Average_hours_of_training_GRI_404_1_formerly_G4_LA9",
        "wikirate_gri_training_hours",
        "Mensch",
        "Qualifizierung",
        "GRI 404-1 average hours of training",
        "hours",
    ),
    (
        "Global_Reporting_Initiative+Incidents_of_Corruption_GRI_Standard_205_3_a",
        "wikirate_gri_corruption_incidents",
        "Demokratie",
        "Korruption/Compliance",
        "GRI 205-3 confirmed incidents of corruption",
        "incidents",
    ),
    (
        "Global_Reporting_Initiative+Environmental_Fines_GRI_307_1_formerly_G4_EN29_a",
        "wikirate_gri_environmental_fines",
        "Demokratie",
        "Rechtskonformität/Umwelt",
        "GRI 307-1 environmental fines",
        "USD",
    ),
]

ALIASES = {
    "BMW": ["BMW Group", "Bayerische Motoren Werke"],
    "Deutsche Post": ["DHL Group", "Deutsche Post DHL Group"],
    "Deutsche Börse": ["Deutsche Boerse"],
    "E.ON": ["E.ON SE", "EON"],
    "Fresenius Medical Care": ["Fresenius Medical Care AG"],
    "GEA Group": ["GEA"],
    "Hannover Re": ["Hannover Rück"],
    "Infineon Technologies": ["Infineon"],
    "Mercedes-Benz Group": ["Mercedes-Benz", "Mercedes-Benz AG"],
    "Merck": ["Merck KGaA"],
    "Munich Re": ["Münchener Rück", "Munich Re Group"],
    "Porsche SE": ["Porsche Automobil Holding SE", "Porsche Automobil Holding"],
    "Qiagen": ["QIAGEN N.V.", "QIAGEN"],
    "ThyssenKrupp": ["thyssenkrupp AG", "thyssenkrupp"],
    "Volkswagen Group": ["Volkswagen", "Volkswagen AG"],
    "1&1 AG": ["1&1", "1&1 Telecom"],
    "Atoss Software AG": ["ATOSS Software"],
    "Cancom SE": ["CANCOM"],
    "Cewe Stiftung & Co. KGaA": ["CEWE Stiftung", "CEWE"],
    "Deutsche Euroshop": ["Deutsche EuroShop"],
}

WIKIRATE_ALIASES = {
    "1&1 AG": ["1&1 Drillisch"],
    "Adtran Networks SE": ["ADVA Optical Networking SE"],
    "Airbus": ["Airbus", "Airbus SE"],
    "Allianz": ["Allianz"],
    "BASF": ["BASF SE"],
    "BMW": ["BMW AG", "Bayerische Motoren Werke AG"],
    "Cancom SE": ["CANCOM"],
    "Covestro": ["Covestro AG"],
    "Deutsche Bank": ["Deutsche Bank AG"],
    "Deutsche Börse": ["Deutsche Börse AG", "Deutsche Boerse AG"],
    "Deutsche Post": ["DHL Group", "Deutsche Post AG", "Deutsche Post DHL Group"],
    "Deutsche Telekom": ["Deutsche Telekom"],
    "E.ON": ["E.ON SE", "E ON SE"],
    "Fresenius": ["Fresenius"],
    "Fresenius Medical Care": ["Fresenius Medical Care AG"],
    "Hannover Re": ["Hannover Rück", "Hannover Re"],
    "Infineon Technologies": ["Infineon Technologies AG"],
    "Mercedes-Benz Group": ["Mercedes-Benz AG"],
    "Merck": ["Merck KGaA"],
    "Munich Re": ["Munich Re", "Münchener Rück"],
    "Porsche SE": ["Porsche Automobil Holding"],
    "Qiagen": ["QIAGEN N.V.", "QIAGEN"],
    "SAP": ["SAP AG", "SAP SE"],
    "Siemens": ["Siemens AG"],
    "ThyssenKrupp": ["thyssenkrupp AG"],
    "Volkswagen Group": ["Volkswagen AG"],
}

BLOCKED_PROVIDER_MATCHES = {
    ("Allianz", "Allianz Investment Management SE"),
    ("BayWa", "Baywa Global Produce"),
    ("Cancom SE", "ScanCom International A/S"),
    ("Fresenius", "Fresenius Medical Care AG"),
    ("Porsche SE", "Porsche Holding GmbH"),
}

PILOT_NAME_MAP = {
    "SAP SE": "SAP",
    "Siemens AG": "Siemens",
    "BASF SE": "BASF",
}


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def normalize(value: str | None) -> str:
    if not value:
        return ""
    text = unicodedata.normalize("NFD", value.casefold())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(
        r"\b(aktiengesellschaft|ag|se|sa|plc|group|holding|holdings|kgaa|gmbh|co|kg|inc|ltd|limited|corporation|corp|company|the|stiftung)\b",
        " ",
        text,
    )
    return re.sub(r"\s+", " ", text).strip()


def stable_id(*parts: Any) -> str:
    return hashlib.sha1("|".join("" if part is None else str(part) for part in parts).encode("utf-8")).hexdigest()[:16]


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def download(url: str, target: Path) -> None:
    request = urllib.request.Request(url, headers={"User-Agent": "wirkungsoekonomie-uwp-impact-importer/0.1"})
    with urllib.request.urlopen(request, timeout=90) as response:
        target.write_bytes(response.read())


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_sbti_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Data"]
    headers = [str(cell) if cell else "" for cell in next(worksheet.iter_rows(values_only=True))]
    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(values_only=True):
        item = dict(zip(headers, row))
        item["_normalized_name"] = normalize(str(item.get("company_name") or ""))
        item["_isin_tokens"] = {token.strip().upper() for token in str(item.get("isin") or "").replace(",", ";").split(";") if token.strip()}
        rows.append(item)
    return rows


def company_candidate_names(company: dict[str, Any]) -> list[str]:
    names = [company["name"], *(ALIASES.get(company["name"], []))]
    return [name for name in names if name]


def find_sbti_match(company: dict[str, Any], rows: list[dict[str, Any]]) -> tuple[dict[str, Any] | None, str, float]:
    company_isin = str(company.get("isin") or "").upper()
    if company_isin:
        isin_candidates = [row for row in rows if company_isin in row.get("_isin_tokens", set())]
        if len(isin_candidates) == 1:
            return isin_candidates[0], "isin_exact_match", 0.98

    normalized_names = {normalize(name) for name in company_candidate_names(company) if normalize(name)}
    candidates: list[tuple[int, dict[str, Any]]] = []
    for row in rows:
        row_name = row["_normalized_name"]
        if not row_name:
            continue
        if row_name in normalized_names:
            candidates.append((5, row))
        elif any(name and (row_name.startswith(name + " ") or row_name.endswith(" " + name)) for name in normalized_names if len(name) >= 4):
            candidates.append((4, row))
        elif any(name and name in row_name for name in normalized_names if len(name) >= 6):
            candidates.append((2, row))
    if not candidates:
        return None, "no_safe_match", 0.0
    candidates.sort(key=lambda item: (item[0], item[1].get("location") == "Germany"), reverse=True)
    selected = candidates[0][1]
    if (company["name"], str(selected.get("company_name") or "")) in BLOCKED_PROVIDER_MATCHES:
        return None, "blocked_subsidiary_or_false_positive", 0.0
    return selected, "safe_name_match", 0.86 if candidates[0][0] >= 4 else 0.68


def normalize_year(value: Any, fallback: int = 2025) -> int:
    if isinstance(value, datetime):
        return value.year
    try:
        if value not in ("", None):
            return int(float(value))
    except (TypeError, ValueError):
        pass
    return fallback


def clean_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if value == "":
        return None
    return value


def wikirate_slug(value: str) -> str:
    text = value.replace("&", "and")
    text = re.sub(r"[^A-Za-z0-9]+", "_", text)
    return text.strip("_")


def fetch_json(url: str, timeout: int = 25) -> dict[str, Any] | None:
    global LAST_FETCH_AT
    if url in FETCH_CACHE:
        return FETCH_CACHE[url]

    elapsed = time.monotonic() - LAST_FETCH_AT
    if elapsed < 0.75:
        time.sleep(0.75 - elapsed)

    request = urllib.request.Request(url, headers={"User-Agent": "wirkungsoekonomie-uwp-impact-importer/0.1", "Accept": "application/json"})
    for attempt in range(4):
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                LAST_FETCH_AT = time.monotonic()
                payload = json.loads(response.read().decode("utf-8"))
                FETCH_CACHE[url] = payload
                return payload
        except urllib.error.HTTPError as error:
            LAST_FETCH_AT = time.monotonic()
            if error.code == 404:
                FETCH_CACHE[url] = None
                return None
            if error.code == 429 and attempt < 3:
                time.sleep(30 + attempt * 30)
                continue
            raise
    FETCH_CACHE[url] = None
    return None


def fetch_text(url: str, timeout: int = 60) -> str | None:
    global LAST_FETCH_AT
    if url in TEXT_FETCH_CACHE:
        return TEXT_FETCH_CACHE[url]

    elapsed = time.monotonic() - LAST_FETCH_AT
    if elapsed < 0.75:
        time.sleep(0.75 - elapsed)

    request = urllib.request.Request(url, headers={"User-Agent": "wirkungsoekonomie-uwp-impact-importer/0.1", "Accept": "text/csv,text/plain,*/*"})
    for attempt in range(4):
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                LAST_FETCH_AT = time.monotonic()
                text = response.read().decode("utf-8", errors="replace")
                TEXT_FETCH_CACHE[url] = text
                return text
        except urllib.error.HTTPError as error:
            LAST_FETCH_AT = time.monotonic()
            if error.code in {403, 404}:
                TEXT_FETCH_CACHE[url] = None
                return None
            if error.code == 429 and attempt < 3:
                time.sleep(30 + attempt * 30)
                continue
            raise
    TEXT_FETCH_CACHE[url] = None
    return None


def wikirate_candidate_names(company: dict[str, Any], sbti_match: dict[str, Any] | None) -> list[str]:
    candidates: list[str] = []

    def add(value: Any) -> None:
        if not value:
            return
        text = str(value).strip()
        if text and text not in candidates:
            candidates.append(text)

    add(company.get("legal_name"))
    add(company.get("name"))
    for alias in ALIASES.get(company["name"], []):
        add(alias)
    for alias in WIKIRATE_ALIASES.get(company["name"], []):
        add(alias)
    if sbti_match:
        add(sbti_match.get("company_name"))

    base = str(company.get("name") or "").strip()
    if base:
        for suffix in ["AG", "SE", "KGaA", "AG & Co. KGaA", "Group", "Holding"]:
            if not base.casefold().endswith(suffix.casefold()):
                add(f"{base} {suffix}")
    return candidates


def build_wikirate_alias_index(
    companies: list[dict[str, Any]],
    company_matches: dict[str, tuple[dict[str, Any] | None, str, float]],
) -> dict[str, dict[str, Any]]:
    alias_map: dict[str, dict[str, Any] | None] = {}
    for company in companies:
        match = company_matches.get(company["company_id"], (None, "", 0.0))[0]
        for candidate in wikirate_candidate_names(company, match):
            key = normalize(candidate)
            if not key:
                continue
            existing = alias_map.get(key)
            if existing and existing["company_id"] != company["company_id"]:
                alias_map[key] = None
            elif key not in alias_map:
                alias_map[key] = company
    return {key: value for key, value in alias_map.items() if value is not None}


def parse_wikirate_csv(text: str) -> list[dict[str, str]]:
    lines = text.splitlines()
    start = next((index for index, line in enumerate(lines) if line.startswith("Answer Page,")), None)
    if start is None:
        return []
    payload = "\n".join(lines[start:])
    if not payload.strip():
        return []
    return list(csv.DictReader(io.StringIO(payload)))


def find_wikirate_company(company: dict[str, Any], sbti_match: dict[str, Any] | None) -> dict[str, Any] | None:
    normalized_candidates = {normalize(name) for name in wikirate_candidate_names(company, sbti_match) if normalize(name)}
    for candidate in wikirate_candidate_names(company, sbti_match):
        slug = wikirate_slug(candidate)
        if not slug:
            continue
        payload = fetch_json(f"{WIKIRATE_BASE_URL}/{slug}.json", timeout=15)
        if not payload:
            continue
        payload_type = payload.get("type", {})
        type_name = payload_type.get("name") if isinstance(payload_type, dict) else payload_type
        if type_name != "Company":
            continue
        provider_name = str(payload.get("name") or candidate)
        if normalize(provider_name) not in normalized_candidates:
            continue
        return {
            "name": provider_name,
            "slug": slug,
            "id": payload.get("id"),
            "url": payload.get("html_url") or f"{WIKIRATE_BASE_URL}/{slug}",
        }
    return None


def parse_wikirate_value(value: Any) -> Any:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.casefold() in {"unknown", "not disclosed", "not found", "n/a", "na", "none"}:
        return None
    normalized = text.replace(",", "").replace("\u202f", "").replace(" ", "")
    try:
        parsed = float(normalized)
        return int(parsed) if parsed.is_integer() else parsed
    except ValueError:
        return text


def make_observation(
    *,
    company: dict[str, Any],
    year: int,
    dimension: str,
    woek_field: str,
    indicator_id: str,
    indicator: str,
    value: Any,
    unit: str,
    data_quality: str,
    source_id: str,
    source_url: str,
    extraction_method: str,
    confidence: float,
    source_anchor: str | None = None,
    note: str | None = None,
) -> dict[str, Any]:
    return {
        "observation_id": f"uwpobs-{stable_id(company['company_id'], indicator_id, year, value, source_id, source_url)}",
        "company_id": company["company_id"],
        "company_name": company["name"],
        "year": year,
        "dimension": dimension,
        "woek_field": woek_field,
        "indicator_id": indicator_id,
        "indicator": indicator,
        "raw_value": clean_value(value),
        "unit": unit,
        "data_quality_class": data_quality,
        "source_id": source_id,
        "source_url": source_url,
        "source_anchor": source_anchor,
        "extraction_method": extraction_method,
        "confidence": round(confidence, 2),
        "score_ready": False,
        "note": note,
    }


def wikirate_observations(company: dict[str, Any], wikirate_company: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not wikirate_company:
        return []

    observations: list[dict[str, Any]] = []
    for metric_slug, indicator_id, dimension, woek_field, indicator, unit in WIKIRATE_METRICS:
        record_url = f"{WIKIRATE_BASE_URL}/{metric_slug}+{wikirate_company['slug']}.json"
        record = fetch_json(record_url, timeout=20)
        if not record:
            continue
        record_items = record.get("items", [])
        if not isinstance(record_items, list):
            continue
        for item in record_items:
            value = parse_wikirate_value(item.get("value"))
            if value is None:
                continue
            year = normalize_year(item.get("year"), fallback=2024)
            source_ids = item.get("sources") or []
            if isinstance(source_ids, list):
                source_anchor = "; ".join(str(source_id) for source_id in source_ids if source_id)
            else:
                source_anchor = str(source_ids)
            comments = item.get("comments")
            if comments:
                source_anchor = f"{source_anchor} | {comments}" if source_anchor else str(comments)
            observations.append(
                make_observation(
                    company=company,
                    year=year,
                    dimension=dimension,
                    woek_field=woek_field,
                    indicator_id=indicator_id,
                    indicator=indicator,
                    value=value,
                    unit=unit,
                    data_quality="C",
                    source_id="wikirate-open-esg",
                    source_url=str(item.get("url") or record_url),
                    source_anchor=source_anchor[:1500] if source_anchor else None,
                    extraction_method="wikirate_public_json_record",
                    confidence=0.62,
                    note=f"WikiRate Community-/Sekundärdaten für {wikirate_company['name']}; Antwortseite und Quellenanker prüfen.",
                )
            )
    return observations


def wikirate_bulk_observations(
    companies: list[dict[str, Any]],
    company_matches: dict[str, tuple[dict[str, Any] | None, str, float]],
) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    alias_index = build_wikirate_alias_index(companies, company_matches)
    observations: list[dict[str, Any]] = []
    matched_names: dict[str, set[str]] = {company["company_id"]: set() for company in companies}
    limit = 1000

    for metric_slug, indicator_id, dimension, woek_field, indicator, unit in WIKIRATE_METRICS:
        offset = 0
        while True:
            csv_url = f"{WIKIRATE_BASE_URL}/{metric_slug}+Answer.csv?limit={limit}&offset={offset}"
            text = fetch_text(csv_url, timeout=60)
            if not text:
                break
            rows = parse_wikirate_csv(text)
            if not rows:
                break
            for row in rows:
                provider_company = (row.get("Company") or "").strip()
                company = alias_index.get(normalize(provider_company))
                if not company:
                    continue
                value = parse_wikirate_value(row.get("Value"))
                if value is None:
                    continue
                matched_names[company["company_id"]].add(provider_company)
                observations.append(
                    make_observation(
                        company=company,
                        year=normalize_year(row.get("Year"), fallback=2024),
                        dimension=dimension,
                        woek_field=woek_field,
                        indicator_id=indicator_id,
                        indicator=indicator,
                        value=value,
                        unit=unit,
                        data_quality="C",
                        source_id="wikirate-open-esg",
                        source_url=str(row.get("Answer Page") or csv_url),
                        source_anchor=str(row.get("Source Page") or "") or None,
                        extraction_method="wikirate_bulk_csv_export",
                        confidence=0.62,
                        note="WikiRate Community-/Sekundärdaten; Antwortseite und Primärquelle prüfen.",
                    )
                )
            if len(rows) < limit:
                break
            offset += limit

    matches: dict[str, dict[str, Any]] = {}
    for company in companies:
        names = sorted(matched_names[company["company_id"]])
        if names:
            matches[company["company_id"]] = {
                "name": ", ".join(names),
                "id": None,
                "url": WIKIRATE_BASE_URL,
            }
    return observations, matches


def sbti_observations(company: dict[str, Any], match: dict[str, Any] | None, match_method: str, confidence: float) -> list[dict[str, Any]]:
    if not match:
        return [
            make_observation(
                company=company,
                year=2025,
                dimension="Planet",
                woek_field="Klima/Transformationspfad",
                indicator_id="sbti_profile_available",
                indicator="SBTi Zielprofil sicher matchbar",
                value=None,
                unit="Status",
                data_quality="E",
                source_id="sbti-target-dashboard",
                source_url=SBTI_URL,
                source_anchor=None,
                extraction_method="official_xlsx_import_no_safe_match",
                confidence=0.0,
                note=f"Kein sicherer SBTi-Match für {company['name']} im offiziellen Target Dashboard. Das ist keine negative Klimabewertung.",
            )
        ]
    observations = []
    for key, dimension, woek_field, indicator, unit, dq in SBTI_FIELDS:
        value = match.get(key)
        if value in ("", None):
            continue
        observations.append(
            make_observation(
                company=company,
                year=normalize_year(match.get("date_updated")),
                dimension=dimension,
                woek_field=woek_field,
                indicator_id=f"sbti_{key}",
                indicator=indicator,
                value=value,
                unit=unit,
                data_quality=dq,
                source_id="sbti-target-dashboard",
                source_url=SBTI_URL,
                source_anchor=str(match.get("company_name") or ""),
                extraction_method=f"official_xlsx_import:{match_method}",
                confidence=confidence,
                note="SBTi-Zieldaten; kein Nachweis tatsächlicher Emissionsreduktion.",
            )
        )
    return observations


def load_pilot(zip_path: Path | None, universe: dict[str, Any]) -> list[dict[str, Any]]:
    if not zip_path or not zip_path.exists():
        return []
    with zipfile.ZipFile(zip_path) as archive:
        payload = json.loads(archive.read("uwp_pilot_wirkungsdaten.json").decode("utf-8"))
    by_name = {company["name"]: company for company in universe["companies"]}
    observations = []
    for pilot_company in payload.get("companies", []):
        target_name = PILOT_NAME_MAP.get(pilot_company.get("name"), pilot_company.get("name"))
        company = by_name.get(target_name)
        if not company:
            continue
        for item in pilot_company.get("datapoints", []):
            indicator_slug = normalize(str(item.get("indicator") or "manual")).replace(" ", "_")[:80] or "manual"
            observations.append(
                make_observation(
                    company=company,
                    year=normalize_year(item.get("year"), fallback=2024),
                    dimension=str(item.get("dimension") or "Datenqualität"),
                    woek_field=str(item.get("woek_field") or "Pilotdaten"),
                    indicator_id=f"pilot_{indicator_slug}",
                    indicator=str(item.get("indicator") or "Pilotdatenpunkt"),
                    value=item.get("value"),
                    unit=str(item.get("unit") or ""),
                    data_quality=str(item.get("data_quality") or "C"),
                    source_id="claude-pilot-public-sources",
                    source_url=str(item.get("source") or ""),
                    source_anchor=None,
                    extraction_method="provided_manual_pilot_import",
                    confidence=0.72 if item.get("value") is not None else 0.4,
                    note=item.get("note"),
                )
            )
    return observations


def gap_observations(company: dict[str, Any], covered_keys: set[str]) -> list[dict[str, Any]]:
    observations = []
    for key, dimension, woek_field, indicator, unit in CORE_GAP_FIELDS:
        if key in covered_keys:
            continue
        observations.append(
            make_observation(
                company=company,
                year=2024,
                dimension=dimension,
                woek_field=woek_field,
                indicator_id=key,
                indicator=indicator,
                value=None,
                unit=unit,
                data_quality="E",
                source_id="uwp-required-field-gap",
                source_url="",
                source_anchor=None,
                extraction_method="required_field_gap_rendered",
                confidence=0.0,
                note="UWP-Kernfeld noch nicht maschinell aus einer versionierten Primärquelle belegt; nicht schätzen, nicht mit 0 füllen.",
            )
        )
    return observations


def covered_gap_keys(observations: list[dict[str, Any]]) -> dict[str, set[str]]:
    mapping: dict[str, set[str]] = {}
    for obs in observations:
        if obs.get("raw_value") is None:
            continue
        cid = obs["company_id"]
        indicator = normalize(str(obs.get("indicator", "")))
        field = normalize(str(obs.get("woek_field", "")))
        keys = mapping.setdefault(cid, set())
        if "scope 1 2" in indicator or ("klima" in field and "absolut" in indicator):
            keys.add("scope_1_2_ghg_absolute")
        if "scope 3" in indicator:
            keys.add("scope_3_ghg_absolute")
        if "taxonomie" in field or "taxonomie" in indicator:
            keys.add("taxonomy_aligned_revenue_share")
        if "erneuerbarer strom" in indicator:
            keys.add("renewable_electricity_share")
        if "frauen" in indicator and ("management" in indicator or "fuhrung" in indicator):
            keys.add("women_management_share")
        if "ltifr" in indicator or "lost time injury" in indicator:
            keys.add("ltifr")
        if "ftse4good" in indicator or "steuertransparenz" in field:
            keys.add("cbcr_tax_transparency")
        if "lobby" in indicator or "lobby" in field:
            keys.add("lobbying_transparency")
        if "whistle" in indicator:
            keys.add("whistleblowing_cases")
        if "datenschutz" in field or "privacy" in indicator:
            keys.add("data_privacy_incidents")
    return mapping


def profile_summary(
    company: dict[str, Any],
    observations: list[dict[str, Any]],
    match: dict[str, Any] | None,
    match_method: str,
    wikirate_match: dict[str, Any] | None,
) -> dict[str, Any]:
    company_observations = [obs for obs in observations if obs["company_id"] == company["company_id"]]
    sourced = [obs for obs in company_observations if obs["data_quality_class"] != "E"]
    gaps = [obs for obs in company_observations if obs["data_quality_class"] == "E"]
    dimensions = sorted({obs["dimension"] for obs in sourced})
    return {
        "company_id": company["company_id"],
        "name": company["name"],
        "ticker": company.get("ticker"),
        "sector": company.get("sector"),
        "sbti_match": {
            "matched": bool(match),
            "method": match_method,
            "provider_name": match.get("company_name") if match else None,
            "provider_sector": match.get("sector") if match else None,
            "provider_location": match.get("location") if match else None,
            "isin": match.get("isin") if match else None,
            "lei": match.get("lei") if match else None,
        },
        "wikirate_match": {
            "matched": bool(wikirate_match),
            "provider_name": wikirate_match.get("name") if wikirate_match else None,
            "provider_id": wikirate_match.get("id") if wikirate_match else None,
            "provider_url": wikirate_match.get("url") if wikirate_match else None,
        },
        "sourced_observation_count": len(sourced),
        "gap_count": len(gaps),
        "dimensions_with_sourced_data": dimensions,
        "score_ready": False,
        "profile_status": "source_bound_partial_profile" if sourced else "metadata_plus_explicit_gaps",
        "interpretation": (
            "Quellengebundenes Teilprofil: echte Beobachtungen und E-Lücken liegen vor. "
            "Kein Gesamtwert, kein Unternehmensranking und keine Investmentempfehlung."
        ),
    }


def write_csv(path: Path, observations: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "observation_id",
        "company_id",
        "company_name",
        "year",
        "dimension",
        "woek_field",
        "indicator_id",
        "indicator",
        "raw_value",
        "unit",
        "data_quality_class",
        "source_id",
        "source_url",
        "source_anchor",
        "extraction_method",
        "confidence",
        "note",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for obs in observations:
            writer.writerow({field: obs.get(field) for field in fields})


def build(args: argparse.Namespace) -> Path:
    universe = read_json(UNIVERSE_PATH)
    companies = universe.get("companies", [])
    if len(companies) != 100:
        raise SystemExit(f"Expected 100 companies, found {len(companies)}")

    with tempfile.TemporaryDirectory() as temp_dir:
        xlsx_path = Path(temp_dir) / "sbti-companies.xlsx"
        download(SBTI_URL, xlsx_path)
        sbti_rows = load_sbti_rows(xlsx_path)
        sbti_hash = file_hash(xlsx_path)

    company_matches: dict[str, tuple[dict[str, Any] | None, str, float]] = {}
    wikirate_matches: dict[str, dict[str, Any] | None] = {}
    observations: list[dict[str, Any]] = []
    for company in companies:
        match, method, confidence = find_sbti_match(company, sbti_rows)
        company_matches[company["company_id"]] = (match, method, confidence)
        observations.extend(sbti_observations(company, match, method, confidence))

    wikirate_observations_all, wikirate_matches_found = wikirate_bulk_observations(companies, company_matches)
    observations.extend(wikirate_observations_all)
    for company in companies:
        wikirate_matches[company["company_id"]] = wikirate_matches_found.get(company["company_id"])

    pilot_observations = load_pilot(args.pilot_zip, universe)
    observations.extend(pilot_observations)

    covered = covered_gap_keys(observations)
    for company in companies:
        observations.extend(gap_observations(company, covered.get(company["company_id"], set())))

    observations.sort(key=lambda item: (item["company_id"], item["data_quality_class"] == "E", item["dimension"], item["indicator_id"]))
    profiles = [
        profile_summary(
            company,
            observations,
            company_matches[company["company_id"]][0],
            company_matches[company["company_id"]][1],
            wikirate_matches[company["company_id"]],
        )
        for company in companies
    ]
    sourced_count = sum(1 for obs in observations if obs["data_quality_class"] != "E")
    gap_count = sum(1 for obs in observations if obs["data_quality_class"] == "E")
    matched_count = sum(1 for match, _, _ in company_matches.values() if match)
    wikirate_matched_count = sum(1 for match in wikirate_matches.values() if match)
    wikirate_observation_count = sum(1 for obs in observations if obs.get("source_id") == "wikirate-open-esg")
    snapshot = {
        "snapshot_id": "uwp-100.impact-data.latest",
        "title": "UWP-100 Wirkungsdaten-Snapshot",
        "universe_id": universe["universe_id"],
        "method_version": METHOD_VERSION,
        "generated_at": utc_now(),
        "score_ready": False,
        "score_note": "Dieser Snapshot enthält belegte Beobachtungen und explizite Datenlücken. Er berechnet keine FinalScores.",
        "company_count": len(companies),
        "source_summary": {
            "sbti_matched_company_count": matched_count,
            "sbti_unmatched_company_count": len(companies) - matched_count,
            "wikirate_matched_company_count": wikirate_matched_count,
            "wikirate_unmatched_company_count": len(companies) - wikirate_matched_count,
            "wikirate_observation_count": wikirate_observation_count,
            "manual_pilot_company_count": len({obs["company_id"] for obs in pilot_observations}),
            "sourced_observation_count": sourced_count,
            "gap_observation_count": gap_count,
            "total_observation_count": len(observations),
        },
        "data_quality_legend": {
            "A": "Primärdaten oder unabhängige Validierung aus öffentlicher Quelle",
            "B": "Primär-/Providerdaten plausibilisiert, aber begrenzter Kontext",
            "C": "Sekundär-/Drittquelle oder externes Rating",
            "D": "Schätzung",
            "E": "Datenlücke; kein Wert",
        },
        "sources": [
            {
                "source_id": "sbti-target-dashboard",
                "title": "Science Based Targets initiative Target Dashboard",
                "url": SBTI_URL,
                "retrieved_at": utc_now(),
                "file_hash_sha256": sbti_hash,
                "used_for": "Planet/Transformation: Zielstatus, Zielklassifikation, Zieljahr, Net-Zero-Status.",
                "limits": "Zieldaten, keine Ist-Emissionen und kein Nachweis tatsächlicher Emissionsreduktion.",
            },
            {
                "source_id": "claude-pilot-public-sources",
                "title": "Gelieferter Pilot-Snapshot mit öffentlichen Einzelquellen",
                "url": str(args.pilot_zip) if args.pilot_zip else None,
                "retrieved_at": utc_now(),
                "used_for": "Zusätzliche manuell recherchierte Einzelwerte für SAP, Siemens, BASF.",
                "limits": "Vom gelieferten ZIP übernommen; Quellen bleiben je Beobachtung erhalten.",
            },
            {
                "source_id": "wikirate-open-esg",
                "title": "WikiRate Open ESG data",
                "url": WIKIRATE_BASE_URL,
                "retrieved_at": utc_now(),
                "used_for": "Öffentliche GRI-Zeitreihen aus WikiRate-Record-JSONs, vor allem Klima-, Arbeitssicherheits-, Trainings-, Abfall- und Compliance-Felder.",
                "limits": "Community-/Sekundärdaten unter CC BY 4.0; DQ C und immer gegen Primärbericht prüfen.",
            },
            {
                "source_id": "uwp-required-field-gap",
                "title": "UWP-Kernfeld als Datenlücke",
                "url": None,
                "retrieved_at": utc_now(),
                "used_for": "Explizite E-Markierung, wenn ein UWP-Kernfeld noch nicht aus einer versionierten Quelle belegt ist.",
                "limits": "Keine Schätzung und keine Bewertung.",
            },
        ],
        "companies": companies,
        "profiles": profiles,
        "observations": observations,
    }

    json_path = SNAPSHOT_DIR / "uwp-100.impact-data.latest.json"
    csv_path = SNAPSHOT_DIR / "uwp-100.impact-data.latest.csv"
    write_json(json_path, snapshot)
    write_csv(csv_path, observations)
    return json_path


def validate() -> None:
    snapshot = read_json(SNAPSHOT_DIR / "uwp-100.impact-data.latest.json")
    if snapshot.get("company_count") != 100:
        raise SystemExit("Snapshot must contain exactly 100 companies")
    if any(profile.get("score_ready") for profile in snapshot.get("profiles", [])):
        raise SystemExit("Profiles must not be score-ready")
    ids = {company["company_id"] for company in snapshot.get("companies", [])}
    for obs in snapshot.get("observations", []):
        if obs.get("company_id") not in ids:
            raise SystemExit(f"Observation for unknown company: {obs.get('company_id')}")
        if obs.get("data_quality_class") not in {"A", "B", "C", "D", "E"}:
            raise SystemExit(f"Invalid DQ: {obs.get('data_quality_class')}")
    print("UWP impact snapshot valid")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("command", nargs="?", choices=["build", "validate"], default="build")
    parser.add_argument("--pilot-zip", type=Path, default=Path("/Users/hagen/Downloads/files.zip"))
    args = parser.parse_args()
    if args.command == "validate":
        validate()
        return
    path = build(args)
    print(path.relative_to(ROOT))


if __name__ == "__main__":
    main()
