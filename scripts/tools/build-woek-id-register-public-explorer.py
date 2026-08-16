#!/usr/bin/env python3
"""Build the public WÖk master-register from the canonical v1.4 workbook."""
from __future__ import annotations

import csv
import hashlib
import html
import json
import re
import shutil
import zipfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
VERSION = "v1.4"
PUBLISHED_AT = "2026-08-16"
PUBLISHER = "Institut für Wirkungsökonomie"
FILENAME = "WOeK_Masterregister_v1.4_FINAL_2026-08-16.xlsx"
SOURCE_XLSX = ROOT / "data" / "master-register" / FILENAME
DOWNLOAD_DIR = ROOT / "assets" / "downloads" / "woek-register" / "v1.4"
CONTENT_DIR = ROOT / "content" / "woek-register"
DATA_PATH = ROOT / "assets" / "data" / "woek-id-register.json"
REGISTER_DIR = ROOT / "woek-id-register"
OVERVIEW_DIR = ROOT / "register"
DATA_ALIAS_DIR = ROOT / "daten" / "woek-register"
LEGACY_TOOL_DIR = ROOT / "werkzeuge" / "woek-id-register"
LIBRARY_DETAIL = ROOT / "bibliothek" / "woek-master-items-register" / "index.html"
HEADER_ROW = 4
PROHIBITED = [
    re.compile(pattern, re.IGNORECASE)
    for pattern in (r"chatgpt", r"openai", r"claude", r"codex", r"anthropic", r"/(?:Users|private|home)/", r"file://", r"sandbox:")
]


def digest(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def esc(value: object) -> str:
    return html.escape("" if value is None else str(value).replace("—", "-"), quote=True)


def slugify(value: object) -> str:
    text = str(value or "").strip().lower()
    for source, target in {"ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss", "–": "-", "‑": "-"}.items():
        text = text.replace(source, target)
    chunks = "".join(character if character.isalnum() else "-" for character in text).split("-")
    return "-".join(chunk for chunk in chunks if chunk)


def records(workbook, name: str) -> list[dict[str, object]]:
    sheet = workbook[name]
    headers = [str(cell.value or "").strip() for cell in sheet[HEADER_ROW]]
    result: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        record = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        if any(value not in (None, "") for value in record.values()):
            result.append(record)
    return result


def multi(value: object) -> list[str]:
    return [part.strip() for part in str(value or "").replace("\n", ";").split(";") if part.strip()]


def german_polarity(value: object) -> str:
    labels = {
        "higher_is_better": "Höherer Messwert ist besser",
        "lower_is_better": "Niedrigerer Messwert ist besser",
        "target_is_better": "Zielwert oder Zielkorridor ist maßgeblich",
        "qualitative": "Qualitative fachliche Einordnung",
    }
    raw = str(value or "").strip()
    return labels.get(raw.lower(), raw or "Nicht festgelegt")


def is_open(item: dict[str, object]) -> bool:
    text = " ".join(str(item.get(key, "")) for key in ["threshold_status", "benchmark_requirement", "source_status", "publication_readiness", "editorial_note"]).lower()
    return any(token in text for token in ["offen", "validieren", "erforderlich", "prüfen", "nicht aktiv"])


def normalize(items_raw, benchmarks_raw, rules_raw, sources_raw):
    benchmarks = {str(row.get("WOK_ID") or ""): row for row in benchmarks_raw}
    rules = {str(row.get("Rule_ID") or ""): row for row in rules_raw}
    source_map = {str(row.get("Source_ID") or ""): row for row in sources_raw}
    result = []
    for row in items_raw:
        identifier = str(row.get("WOK_ID") or "").strip()
        if not identifier:
            continue
        rule = rules.get(str(row.get("Rule_ID") or ""), {})
        benchmark = benchmarks.get(identifier, {})
        source_ids = multi(row.get("Source_IDs", ""))
        item = {
            "id": identifier, "slug": slugify(identifier),
            "sdg_or_sdgplus": row.get("SDG_or_SDGplus", ""), "target": row.get("Target/Unterziel", ""),
            "indicator_family": row.get("Indikatorfamilie", ""), "item": row.get("Item", ""),
            "definition": row.get("Definition/Messgröße", ""), "unit": row.get("Einheit", ""),
            "polarity": row.get("Polarity", ""), "polarity_label": german_polarity(row.get("Polarity", "")),
            "rule_id": row.get("Rule_ID", ""), "rule_type": rule.get("Regeltyp", ""), "input_mode": rule.get("Eingabemodus", ""),
            "thresholds": row.get("Schwellen (WÖk-Klassen)", ""), "threshold_category": row.get("Schwellenkategorie", ""),
            "threshold_status": row.get("Schwellenstatus", ""), "threshold_basis": row.get("Grenzwertbasis", ""),
            "benchmark_requirement": row.get("Benchmarkbedarf", ""), "benchmark_status": benchmark.get("Active_Status", ""),
            "source_ids": source_ids, "source_detail": row.get("Quelle_detail", ""),
            "source_function": row.get("Quellenfunktion", ""), "source_status": row.get("Quellenstatus", ""),
            "nace_legacy": row.get("NACE_Beispiele_Legacy", ""), "nace_version": row.get("NACE_Version", ""),
            "nace_rev21": row.get("NACE_Rev2.1_Beispiele", ""), "nace_status": row.get("NACE_Status", ""),
            "calculation": row.get("Berechnungslogik", ""), "system_boundary": row.get("Systemgrenze", ""),
            "data_quality_minimum": row.get("Datenqualitätsanforderung", ""), "assurance_level_required": row.get("Assurance_Anforderung", ""),
            "publication_readiness": row.get("Fachlogik_Status", ""), "audit_priority": row.get("Prüfpriorität", ""),
            "editorial_note": row.get("Prüfhinweis", ""), "version": row.get("Version", ""),
            "valid_from": row.get("Gültig_ab", ""), "valid_to": row.get("Gültig_bis", ""),
            "mpd_note": "Die Zuordnung zu Mensch, Planet und Demokratie wird im konkreten Wirkpfad vorgenommen und nicht pauschal aus dem SDG abgeleitet.",
            "sources": [{
                "id": source_id, "slug": slugify(source_id),
                "organisation": source_map.get(source_id, {}).get("Organisation", ""),
                "title": source_map.get(source_id, {}).get("Titel/Standard", source_id),
                "function": source_map.get(source_id, {}).get("Funktion", ""),
                "version": source_map.get(source_id, {}).get("Version/Stand", ""),
                "url": source_map.get(source_id, {}).get("URL", ""),
                "notes": source_map.get(source_id, {}).get("Hinweise", ""),
            } for source_id in source_ids],
        }
        item["open_status"] = is_open(item)
        result.append(item)
    return result


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")


def layout(title: str, description: str, body: str, prefix: str, canonical_path: str) -> str:
    return f'''<!doctype html>
<html lang="de"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>{esc(title)}</title><meta name="description" content="{esc(description)}">
<meta name="search_title" content="{esc(title)}"><meta name="search_description" content="{esc(description)}"><meta name="search_section" content="WÖk-Masterregister"><meta name="search_type" content="Register">
<link rel="stylesheet" href="{prefix}assets/css/style.css"><link rel="canonical" href="https://wirkungsoekonomie.de{esc(canonical_path)}"><link rel="icon" href="{prefix}assets/img/brand/favicon.svg" type="image/svg+xml"></head>
<body><header class="site-header" data-search-exclude><a class="brand" href="{prefix}index.html" aria-label="Wirkungsökonomie Startseite"><span class="brand-mark"><img src="{prefix}assets/img/brand/signet.svg" alt="Wirkungsökonomie Logo"></span><span class="brand-name">Wirkungsökonomie</span></a><button class="nav-toggle" type="button" aria-label="Menü öffnen" aria-expanded="false" aria-controls="site-nav"><span class="nav-toggle-icon" aria-hidden="true">☰</span><span class="sr-only">Menü</span></button><nav class="site-nav" id="site-nav" aria-label="Hauptnavigation" data-search-exclude><a href="{prefix}index.html">Start</a><a href="{prefix}verstehen/">Verstehen</a><a href="{prefix}wirkungsfelder/">Wirkungsfelder</a><a href="{prefix}werkzeuge/">Praxis &amp; Tools</a><a href="{prefix}bibliothek/">Bibliothek</a><a href="{prefix}suche.html">Suche</a></nav></header><main data-pagefind-body>{body}</main><script src="{prefix}assets/js/main.js"></script></body></html>'''


def overview_page(items: list[dict[str, object]]) -> str:
    body = f'''<section class="hero portal-hero"><div class="hero-content"><nav class="breadcrumb"><a href="../index.html">Start</a> / Register</nav><p class="hero-kicker">Kanonische technische Registerfassung {VERSION}</p><h1>WÖk-Masterregister</h1><p class="hero-subtitle">Die technische Grundlage für WÖk-IDs, Messregeln, Quellen, Benchmarks und Kalibrierungsstatus.</p><p>Das Register dokumentiert {len(items)} Wirkungsindikatoren. Es zeigt nicht nur nutzbare Regeln, sondern auch offene Kalibrierungen, fehlende Benchmarks und Prüfbedarf.</p><p class="callout"><strong>Wichtig:</strong> Fehlende Daten bedeuten keine neutrale Bewertung. Sie bleiben unbewertet.</p><div class="hero-actions"><a class="btn btn-primary" href="../woek-id-register/">Register durchsuchen</a><a class="btn btn-secondary" href="../bibliothek/woek-master-items-register/">Registerbeschreibung und Downloads</a></div></div></section>'''
    return layout("WÖk-Masterregister", "Kanonisches WÖk-Masterregister v1.4 mit WÖk-IDs, Messregeln, Quellen, Benchmarks und sichtbarem Prüfstatus.", body, "../", "/register/")


def register_page(items, rules, audit, changelog) -> str:
    embedded = json.dumps({"items": items}, ensure_ascii=False, default=str).replace("<", "\\u003c")
    rule_options = "".join(f'<option value="{esc(rule.get("Rule_ID"))}">{esc(rule.get("Rule_ID"))}</option>' for rule in rules if rule.get("Rule_ID"))
    status_options = "".join(f'<option value="{esc(status)}">{esc(status)}</option>' for status in sorted({str(item["publication_readiness"]) for item in items if item["publication_readiness"]}))
    sdg_options = "".join(f'<option value="{esc(status)}">{esc(status)}</option>' for status in sorted({str(item["sdg_or_sdgplus"]) for item in items if item["sdg_or_sdgplus"]}))
    body = f'''<section class="hero compact-hero"><nav class="breadcrumb"><a href="../index.html">Start</a> / <a href="../register/">Register</a></nav><p class="hero-kicker">Kanonische Fassung {VERSION} · 16. August 2026</p><h1>WÖk-Masterregister durchsuchen</h1><p class="hero-subtitle">{len(items)} WÖk-IDs mit Definition, Messrichtung, Regel, Quellenfunktion, Schwellenstatus, Datenqualität und Prüfbedarf.</p><p class="callout"><strong>FINAL bezeichnet die führende Registerfassung.</strong> Das bedeutet nicht, dass jede Schwelle extern validiert ist. Offene Kalibrierungen und Benchmarks bleiben sichtbar.</p></section>
<section class="section"><div class="section-header"><h2>Indikatoren und Regeln finden</h2><p>Die Detailseiten erklären, was gemessen wird, welche Quelle welche Funktion erfüllt und wo noch fachliche Prüfung nötig ist.</p></div><form class="woek-register-filters" data-search-exclude><label>Suche<input id="woekSearch" type="search" placeholder="WOK-S-101, Lebenslohn, Wasser"></label><label>SDG oder SDG+<select id="filterSdg"><option value="">Alle</option>{sdg_options}</select></label><label>Regel<select id="filterRule"><option value="">Alle</option>{rule_options}</select></label><label>Fachstatus<select id="filterStatus"><option value="">Alle</option>{status_options}</select></label><label>Offene Prüfung<select id="filterOpen"><option value="">Alle</option><option value="true">Offene Validierung oder Benchmarkfrage</option><option value="false">Strukturell nutzbar</option></select></label></form><p class="text-note" id="registerCount" aria-live="polite"></p><div class="table-wrap"><table class="data-table"><thead><tr><th>WÖk-ID</th><th>Indikator</th><th>SDG / SDG+</th><th>Messrichtung</th><th>Regel</th><th>Fachstatus</th><th>Details</th></tr></thead><tbody id="registerRows"></tbody></table></div></section>
<section class="section"><div class="card-grid three"><article class="card"><h2 class="card-title">Daten und Downloads</h2><p>XLSX, CSV, JSON, Version und Prüfsummen stehen auf einer vorgeschalteten Dokumentseite.</p><a class="text-link" href="../bibliothek/woek-master-items-register/">Registerbeschreibung öffnen</a></article><article class="card"><h2 class="card-title">Prüfprotokoll</h2><p>{len(audit)} dokumentierte Prüf- und Korrekturpunkte im Arbeitsstand.</p><a class="text-link" href="methodik/">Methodik und Grenzen lesen</a></article><article class="card"><h2 class="card-title">Versionierung</h2><p>{len(changelog)} dokumentierte Versionsschritte. v1.4 ersetzt v1.3 als führende technische Registerquelle.</p><a class="text-link" href="../dokumente/woek-master-items-final-v1-2/">Historische v1.2 ansehen</a></article></div></section>
<script id="woekRegisterData" type="application/json">{embedded}</script><script>const data=JSON.parse(document.getElementById('woekRegisterData').textContent),items=data.items,$=id=>document.getElementById(id),safe=v=>String(v??'').replace(/[&<>"']/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[c]));function render(){{const q=$('woekSearch').value.toLowerCase().trim(),sdg=$('filterSdg').value,rule=$('filterRule').value,status=$('filterStatus').value,open=$('filterOpen').value,rows=items.filter(i=>(!q||[i.id,i.item,i.definition,i.sdg_or_sdgplus,i.indicator_family,i.source_detail,i.source_function].join(' ').toLowerCase().includes(q))&&(!sdg||i.sdg_or_sdgplus===sdg)&&(!rule||i.rule_id===rule)&&(!status||i.publication_readiness===status)&&(!open||String(i.open_status)===open));$('registerCount').textContent=`${{rows.length}} von ${{items.length}} WÖk-IDs sichtbar`;$('registerRows').innerHTML=rows.slice(0,250).map(i=>`<tr><td><strong>${{safe(i.id)}}</strong></td><td>${{safe(i.item)}}</td><td>${{safe(i.sdg_or_sdgplus)}}</td><td>${{safe(i.polarity_label)}}</td><td><code>${{safe(i.rule_id)}}</code></td><td>${{safe(i.publication_readiness)}}</td><td><a class="btn btn-secondary table-action" href="${{safe(i.slug)}}/">Eintrag prüfen</a></td></tr>`).join('')+(rows.length>250?'<tr><td colspan="7">Weitere Treffer vorhanden. Bitte Filter verfeinern.</td></tr>':'');}}document.querySelectorAll('.woek-register-filters input,.woek-register-filters select').forEach(element=>element.addEventListener('input',render));render();</script>'''
    return layout("WÖk-Masterregister v1.4", "Durchsuchbares WÖk-Masterregister v1.4 mit 621 Wirkungsindikatoren, Regeln, Quellen, Kalibrierungen und sichtbarem Prüfstatus.", body, "../", "/woek-id-register/")


def detail_page(item: dict[str, object]) -> str:
    source_rows = "".join(f'<li><a href="../quellen/{esc(source["slug"])}/"><strong>{esc(source["id"])}:</strong> {esc(source["title"])}</a><br><span class="text-note">Funktion: {esc(source["function"])}</span></li>' for source in item["sources"]) or "<li>Keine Quelle hinterlegt.</li>"
    status = "Fachliche Validierung, Quelle oder Benchmark noch offen" if item["open_status"] else "Strukturell nutzbarer Registereintrag"
    body = f'''<section class="hero compact-hero"><nav class="breadcrumb"><a href="../../register/">Register</a> / <a href="../">WÖk-Masterregister</a> / {esc(item["id"])}</nav><p class="hero-kicker">WÖk-Masterregister {VERSION}</p><h1>{esc(item["item"])}</h1><p class="hero-subtitle"><strong>{esc(item["id"])}</strong> · {esc(item["definition"])}</p><p class="callout"><strong>{esc(status)}.</strong> {esc(item["publication_readiness"])}</p></section><section class="section document-detail-grid"><article class="document-detail-main"><h2>Was wird gemessen?</h2><p>{esc(item["definition"])}</p><dl><dt>Einheit</dt><dd>{esc(item["unit"])}</dd><dt>Technische Messrichtung</dt><dd>{esc(item["polarity_label"])}</dd></dl><h2>Zielbezug</h2><p>{esc(item["sdg_or_sdgplus"])} · Unterziel oder Referenzbereich {esc(item["target"])}</p><p class="text-note">{esc(item["mpd_note"])}</p><h2>Regel und Schwellen</h2><p><strong>Regel-ID:</strong> <code>{esc(item["rule_id"])}</code> · {esc(item["rule_type"])} · {esc(item["input_mode"])}</p><p><strong>Schwellen:</strong> {esc(item["thresholds"])}</p><p><strong>Art und Status:</strong> {esc(item["threshold_category"])} · {esc(item["threshold_status"])}</p><p><strong>Grenzwertbasis:</strong> {esc(item["threshold_basis"])}</p><h2>Benchmark</h2><p>{esc(item["benchmark_requirement"])}{(' · ' + esc(item["benchmark_status"])) if item["benchmark_status"] else ''}</p><h2>Berechnung und Systemgrenze</h2><p><strong>Berechnungslogik:</strong> {esc(item["calculation"])}</p><p><strong>Systemgrenze:</strong> {esc(item["system_boundary"])}</p><h2>Quellen und ihre Funktion</h2><p>{esc(item["source_detail"])}</p><p><strong>Quellenfunktion:</strong> {esc(item["source_function"])}</p><p><strong>Quellenstatus:</strong> {esc(item["source_status"])}</p><ul>{source_rows}</ul><h2>Branchenbezug</h2><p>{esc(item["nace_legacy"] or "Kein historisches NACE-Beispiel hinterlegt.")}</p><p>{esc(item["nace_status"])}</p></article><aside class="document-detail-aside" data-search-exclude><dl><dt>Datenqualität</dt><dd>{esc(item["data_quality_minimum"])}</dd><dt>Prüftiefe</dt><dd>{esc(item["assurance_level_required"])}</dd><dt>Fachstatus</dt><dd>{esc(item["publication_readiness"])}</dd><dt>Prüfpriorität</dt><dd>{esc(item["audit_priority"])}</dd><dt>Offene Fachfrage</dt><dd>{esc(item["editorial_note"])}</dd><dt>Version</dt><dd>{esc(item["version"])}</dd><dt>Gültig ab</dt><dd>{esc(item["valid_from"])}</dd></dl><a class="btn btn-secondary" href="../">Zurück zum Register</a></aside></section>'''
    return layout(f'{item["id"]} · {item["item"]}', f'Detailseite zur WÖk-ID {item["id"]} mit Messgröße, Regel, Quellenfunktion, Kalibrierungs- und Prüfstatus.', body, "../../", f'/woek-id-register/{item["slug"]}/')


def source_index_page(sources) -> str:
    cards = "".join(f'<article class="card"><p class="card-kicker">{esc(source.get("Source_ID"))}</p><h2 class="card-title">{esc(source.get("Titel/Standard"))}</h2><p>{esc(source.get("Organisation"))}</p><p><strong>Funktion:</strong> {esc(source.get("Funktion"))}</p><a class="text-link" href="{slugify(source.get("Source_ID"))}/">Quelle einordnen</a></article>' for source in sources)
    body = f'''<section class="hero compact-hero"><nav class="breadcrumb"><a href="../../register/">Register</a> / <a href="../">WÖk-Masterregister</a></nav><p class="hero-kicker">Quellenkatalog {VERSION}</p><h1>Welche Quellen erfüllen welche Funktion?</h1><p class="hero-subtitle">Messstandard, Zielrahmen, Rechtsgrundlage, Datensatz und WÖk-Kalibrierung sind nicht dasselbe. Jede Quelle wird deshalb zuerst eingeordnet.</p></section><section class="section"><div class="card-grid three">{cards}</div></section>'''
    return layout("Quellen des WÖk-Masterregisters", "Interne Detailseiten zu den Quellen des WÖk-Masterregisters v1.4 mit Herausgeber, Funktion, Version und Originalzugang.", body, "../../", "/woek-id-register/quellen/")


def source_detail_page(source: dict[str, object]) -> str:
    url = str(source.get("URL") or "").strip()
    original = f'<a class="btn btn-primary" href="{esc(url)}" rel="noopener noreferrer">Originalquelle öffnen</a>' if url else '<p class="callout">Für diese interne Referenz ist kein externer Originalzugang hinterlegt.</p>'
    body = f'''<section class="hero compact-hero"><nav class="breadcrumb"><a href="../../../register/">Register</a> / <a href="../../">WÖk-Masterregister</a> / <a href="../">Quellen</a></nav><p class="hero-kicker">Quellendetail · {esc(source.get("Source_ID"))}</p><h1>{esc(source.get("Titel/Standard"))}</h1><p class="hero-subtitle">{esc(source.get("Organisation"))}</p></section><section class="section document-detail-grid"><article class="document-detail-main"><h2>Welche Rolle spielt diese Quelle?</h2><p>{esc(source.get("Funktion"))}</p><h2>Fassung und Stand</h2><p>{esc(source.get("Version/Stand") or "Kein Versionsstand hinterlegt.")}</p><h2>Einordnung und Grenzen</h2><p>{esc(source.get("Hinweise") or "Keine zusätzliche Einordnung hinterlegt.")}</p><p>Eine Mess- oder Berichtsquelle begründet nicht automatisch die Schwellen einer WÖk-Bewertung. Die konkrete Quellenfunktion und der Schwellenstatus bleiben deshalb im jeweiligen Registereintrag sichtbar.</p></article><aside class="document-detail-aside" data-search-exclude><dl><dt>Quellen-ID</dt><dd>{esc(source.get("Source_ID"))}</dd><dt>Organisation</dt><dd>{esc(source.get("Organisation"))}</dd><dt>Funktion</dt><dd>{esc(source.get("Funktion"))}</dd><dt>Stand</dt><dd>{esc(source.get("Version/Stand"))}</dd></dl>{original}<a class="text-link" href="../">Zur Quellenübersicht</a></aside></section>'''
    return layout(f'{source.get("Titel/Standard")} · Quelle', f'Einordnung der Quelle {source.get("Titel/Standard")} im WÖk-Masterregister.', body, "../../../", f'/woek-id-register/quellen/{slugify(source.get("Source_ID"))}/')


def methodology_page() -> str:
    body = '''<section class="hero compact-hero"><nav class="breadcrumb"><a href="../../register/">Register</a> / <a href="../">WÖk-Masterregister</a></nav><p class="hero-kicker">Methodik und Grenzen</p><h1>Was das Masterregister leistet - und was nicht</h1><p class="hero-subtitle">Das Register strukturiert messbare Wirkungsaspekte. Es ersetzt weder den Wirkpfad eines konkreten Falls noch die fachliche Prüfung.</p></section><section class="section narrow"><div class="card"><h2>Vom Fall zum Indikator</h2><p>Zuerst werden Originalquelle, Wirkmechanismus, mögliche Zustandsveränderung und Referenzziel bestimmt. Erst danach kann eine passende WÖk-ID als Messindikator gewählt werden.</p><h2>Messrichtung ist nicht politische Wirkungsrichtung</h2><p>Die Angabe „höher ist besser“ beschreibt die technische Logik eines Indikators. Ob eine konkrete politische Entscheidung diesen Indikator positiv oder negativ beeinflusst, muss im Wirkpfad begründet werden.</p><h2>Keine Daten bedeuten keine Bewertung</h2><p>Ein fehlender Messwert wird nicht als null und nicht als neutral behandelt. Er bleibt unbewertet.</p><h2>Schwellen bleiben prüfbar</h2><p>Externe Schwellen und WÖk-Kalibrierungen werden getrennt. FINAL bezeichnet die führende Registerfassung, nicht die externe Validierung jeder einzelnen Grenze.</p><h2>Nichtkompensation</h2><p>Schwere negative Wirkungen oder verletzte Schutzgrenzen dürfen nicht durch gute Werte an anderer Stelle weggemittelt werden.</p></div></section>'''
    return layout("Methodik des WÖk-Masterregisters", "Wie WÖk-IDs, Messrichtungen, Schwellen, Quellenfunktionen, Datenlücken und Nichtkompensation im Masterregister getrennt werden.", body, "../../", "/woek-id-register/methodik/")


def library_detail_page(statistics: dict[str, int]) -> str:
    body = f'''<section class="hero compact-hero document-detail-hero"><nav class="breadcrumb"><a href="../../bibliothek/">Bibliothek</a> / Datenregister</nav><p class="hero-kicker">Datenregister · führende Fassung</p><h1>WÖk-Masterregister v1.4</h1><p class="hero-subtitle">Kanonische technische Grundlage für WÖk-IDs, Messregeln, Quellen, Benchmarks, Kalibrierungen und Prüfstatus.</p><div class="document-card-badges"><span class="status-badge">621 WÖk-IDs</span><span class="status-badge">Version 1.4</span><span class="status-badge">Herausgeber: Institut für Wirkungsökonomie</span></div></section><section class="section document-detail-grid"><aside class="document-detail-aside" data-search-exclude><dl><dt>Dokumentart</dt><dd>Datenregister</dd><dt>Status</dt><dd>führende technische Registerfassung</dd><dt>Stand</dt><dd>16. August 2026</dd><dt>Umfang</dt><dd>{statistics["woek_ids"]} WÖk-IDs · {statistics["indicator_families"]} Indikatorfamilien · {statistics["scoring_rules"]} Regeln</dd><dt>Herausgeber</dt><dd>{PUBLISHER}</dd></dl><div class="document-action-row"><a class="btn btn-secondary" href="../../woek-id-register/">Online-Register öffnen</a><a class="btn btn-primary" href="../../assets/downloads/woek-register/v1.4/{FILENAME}">XLSX herunterladen</a><a class="btn btn-secondary" href="../../assets/downloads/woek-register/v1.4/register-v1.4.csv">CSV herunterladen</a><a class="btn btn-secondary" href="../../assets/downloads/woek-register/v1.4/register-v1.4.json">JSON herunterladen</a></div><a class="text-link" href="../../assets/downloads/woek-register/v1.4/manifest.json">Prüfsummen und Manifest</a></aside><article class="document-detail-main"><div class="callout"><strong>Kurz gesagt:</strong> Das Register zeigt, welche Wirkungsaspekte messbar gemacht werden, welche Regeln und Quellen gelten und wo Kalibrierungen, Benchmarks oder Fachprüfungen noch offen sind.</div><h2>Was ist neu in v1.4?</h2><ul><li>Der führende Begriffsleitfaden v1.5 ist als Methodenreferenz verankert.</li><li>SDG+ ist durchgängig als WÖk-Erweiterung gekennzeichnet.</li><li>Quellenfunktion, Schwellenstatus und Benchmarkbedarf bleiben getrennt sichtbar.</li><li>Offene Kalibrierungen wurden nicht künstlich als validiert ausgegeben.</li><li>Fehlende Daten erzeugen keine neutrale Bewertung und keinen Score.</li></ul><h2>Was kann ich prüfen?</h2><p>Jede WÖk-ID hat eine eigene Detailseite. Dort stehen Messgröße, Einheit, technische Messrichtung, Scoring-Regel, Schwellenherkunft, Quellenfunktion, Datenqualitätsanforderung und fachlicher Prüfstatus.</p><h2>Wichtige Grenze</h2><p>Das Register ist keine Parteibewertung und keine automatische politische Entscheidung. Eine WÖk-ID wird erst dann sinnvoll, wenn ein konkreter Wirkpfad zeigt, welche Zustandsveränderung bei wem, wo und unter welchen Bedingungen erwartet oder beobachtet wird.</p><h2>Versionierung</h2><p>v1.4 ersetzt v1.3 als führende technische Registerquelle. Historische Fassungen bleiben als frühere Arbeitsstände erkennbar; neue Anwendungen müssen v1.4 verwenden.</p><p><a class="text-link" href="../../woek-id-register/methodik/">Methodische Grenzen lesen</a> · <a class="text-link" href="../../woek-id-register/quellen/">Quellenfunktionen prüfen</a></p></article></section>'''
    return layout("WÖk-Masterregister v1.4 | Bibliothek der Wirkungsökonomie", "Führendes WÖk-Masterregister v1.4 mit 621 WÖk-IDs, 204 Indikatorfamilien, 28 Regeln, Quellen-, Kalibrierungs- und Prüfstatus.", body, "../../", "/bibliothek/woek-master-items-register/")


def build_exports(items_raw, schema_raw, statistics) -> dict[str, object]:
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    public_xlsx = DOWNLOAD_DIR / FILENAME
    shutil.copyfile(SOURCE_XLSX, public_xlsx)
    fields = [str(row.get("Public_Field") or "") for row in schema_raw if row.get("Public_Field")]
    sources = {str(row.get("Public_Field") or ""): str(row.get("Source_Column") or "") for row in schema_raw if row.get("Public_Field")}
    public_records = [{field: row.get(sources[field], "") for field in fields} for row in items_raw]
    csv_path = DOWNLOAD_DIR / "register-v1.4.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader(); writer.writerows(public_records)
    json_path = DOWNLOAD_DIR / "register-v1.4.json"
    write_json(json_path, {"schema_version": "1.0.0", "register_version": "1.4", "published_at": PUBLISHED_AT, "publisher": PUBLISHER, "status": "kanonische technische Registerfassung", "interpretation_boundary": "FINAL bezeichnet die führende Registerfassung. Offene Kalibrierungen, Benchmarks und Fachprüfungen bleiben sichtbar. Fehlende Daten erzeugen keine Bewertung.", "statistics": statistics, "public_schema": schema_raw, "items": public_records})
    manifest = {"manifest_version": "1.0.0", "register_version": "1.4", "published_at": PUBLISHED_AT, "publisher": PUBLISHER, "canonical_source": FILENAME, "canonical_document_page": "/bibliothek/woek-master-items-register/", "files": [{"filename": path.name, "sha256": digest(path), "bytes": path.stat().st_size} for path in [public_xlsx, csv_path, json_path]], "statistics": statistics, "changelog": "v1.4 ersetzt v1.3 als führende technische Registerquelle. Offene Kalibrierungen und Benchmarks wurden nicht künstlich geschlossen."}
    write_json(DOWNLOAD_DIR / "manifest.json", manifest)
    return manifest


def ensure_public_safety(paths: list[Path]) -> None:
    for path in paths:
        if path.suffix.lower() == ".xlsx":
            with zipfile.ZipFile(path) as archive:
                text = "\n".join(archive.read(name).decode("utf-8", "ignore") for name in archive.namelist() if name.endswith((".xml", ".rels")))
        else:
            text = path.read_text(encoding="utf-8", errors="ignore")
        for pattern in PROHIBITED:
            if match := pattern.search(text):
                raise SystemExit(f"Nicht veröffentlichbare Herkunftsspur in {path}: {match.group(0)}")


def redirect_page(target: str, title: str) -> str:
    return f'<!doctype html><html lang="de"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{esc(title)}</title><meta http-equiv="refresh" content="0; url={esc(target)}"><link rel="canonical" href="https://wirkungsoekonomie.de/woek-id-register/"></head><body><main><h1>{esc(title)}</h1><p><a href="{esc(target)}">Zum WÖk-Masterregister</a></p></main></body></html>'


def main() -> None:
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"Führendes Register fehlt: {SOURCE_XLSX}")
    workbook = load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    raw_items = records(workbook, "01_Item_Register"); raw_benchmarks = records(workbook, "02_Benchmarks")
    raw_rules = records(workbook, "03_Scoring_Rules"); raw_sources = records(workbook, "07_Quellenkatalog")
    audit = records(workbook, "08_Prüfprotokoll"); changelog = records(workbook, "05_Changelog")
    schema = records(workbook, "10_Public_Export_Schema")
    items = normalize(raw_items, raw_benchmarks, raw_rules, raw_sources)
    if len(items) != 621 or len(raw_rules) != 28:
        raise SystemExit(f"Unerwarteter v1.4-Import: {len(items)} Einträge / {len(raw_rules)} Regeln")
    statistics = {"woek_ids": len(items), "indicator_families": len({str(item["indicator_family"]) for item in items if item["indicator_family"]}), "scoring_rules": len(raw_rules), "sdg_plus_assignments": sum(1 for item in items if str(item["sdg_or_sdgplus"]).startswith("SDG+"))}
    manifest = build_exports(raw_items, schema, statistics)
    payload = {"version": VERSION, "published_at": PUBLISHED_AT, "publisher": PUBLISHER, "source_hash_sha256": digest(SOURCE_XLSX), "generated_at": f"{PUBLISHED_AT}T00:00:00+02:00", "statistics": statistics, "items": items, "sources": raw_sources, "methods": raw_rules, "audit": audit, "changelog": changelog, "manifest": manifest}
    for directory in [CONTENT_DIR, REGISTER_DIR, OVERVIEW_DIR, LEGACY_TOOL_DIR, DATA_ALIAS_DIR, LIBRARY_DETAIL.parent]: directory.mkdir(parents=True, exist_ok=True)
    write_json(CONTENT_DIR / "items.json", items); write_json(CONTENT_DIR / "sources.json", raw_sources)
    write_json(CONTENT_DIR / "methods.json", raw_rules); write_json(CONTENT_DIR / "audit-findings.json", audit)
    write_json(CONTENT_DIR / "changelog.json", changelog); write_json(DATA_PATH, payload)
    (OVERVIEW_DIR / "index.html").write_text(overview_page(items), encoding="utf-8")
    (REGISTER_DIR / "index.html").write_text(register_page(items, raw_rules, audit, changelog), encoding="utf-8")
    (REGISTER_DIR / "methodik").mkdir(exist_ok=True); (REGISTER_DIR / "methodik" / "index.html").write_text(methodology_page(), encoding="utf-8")
    source_root = REGISTER_DIR / "quellen"; source_root.mkdir(exist_ok=True); (source_root / "index.html").write_text(source_index_page(raw_sources), encoding="utf-8")
    for source in raw_sources:
        directory = source_root / slugify(source.get("Source_ID")); directory.mkdir(exist_ok=True)
        (directory / "index.html").write_text(source_detail_page(source), encoding="utf-8")
    for item in items:
        directory = REGISTER_DIR / item["slug"]; directory.mkdir(exist_ok=True)
        (directory / "index.html").write_text(detail_page(item), encoding="utf-8")
    LIBRARY_DETAIL.write_text(library_detail_page(statistics), encoding="utf-8")
    (DATA_ALIAS_DIR / "index.html").write_text(redirect_page("../../woek-id-register/", "WÖk-Masterregister"), encoding="utf-8")
    (LEGACY_TOOL_DIR / "index.html").write_text(redirect_page("../../woek-id-register/", "WÖk-Masterregister"), encoding="utf-8")
    ensure_public_safety([SOURCE_XLSX, DOWNLOAD_DIR / FILENAME, DOWNLOAD_DIR / "register-v1.4.csv", DOWNLOAD_DIR / "register-v1.4.json", DOWNLOAD_DIR / "manifest.json"])
    print(f"WÖk-Masterregister {VERSION}: {len(items)} WÖk-IDs, {len(raw_rules)} Regeln, {len(raw_sources)} Quellen veröffentlicht.")


if __name__ == "__main__":
    main()
